#!/usr/bin/env python3
"""
Excel to Lua Converter
Can be called from Lua using os.execute() or io.popen()
Usage: python excel_to_lua.py <input.xlsx> [output.lua]
"""

import sys
import os
import openpyxl

def extract_metadata_from_lua(lua_path):
    """Extract metadata section from existing Lua file if it exists"""
    if not os.path.exists(lua_path):
        return None
    
    try:
        with open(lua_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # Look for the metadata marker
        metadata_marker = "-- METADATA_START"
        if metadata_marker in content:
            # Extract everything from the marker onwards
            metadata_start = content.find(metadata_marker)
            metadata_section = content[metadata_start:]
            return metadata_section
        
        return None
    except Exception as e:
        print(f"WARNING: Could not read existing metadata: {e}", file=sys.stderr)
        return None

def excel_to_lua(excel_path, output_path=None):
    """Convert Excel file to Lua ordered nested tables"""
    
    if output_path is None:
        output_path = excel_path.rsplit('.', 1)[0] + '.lua'
    
    # Preserve existing metadata if file exists
    existing_metadata = extract_metadata_from_lua(output_path)
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        lua_code = []
        lua_code.append("-- Auto-generated Lua table from Excel file")
        lua_code.append(f"-- Source: {excel_path}")
        lua_code.append("-- Structure: workbook[sheet_name][row][column] = value\n")
        lua_code.append("local workbook = {}\n")
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            lua_code.append(f"-- Sheet: {sheet_name}")
            lua_code.append(f'workbook["{sheet_name}"] = {{}}\n')
            
            max_row = ws.max_row
            max_col = ws.max_column
            
            # Process each row
            for row_idx in range(1, max_row + 1):
                lua_code.append(f'workbook["{sheet_name}"][{row_idx}] = {{}}')
                
                # Process each column in the row
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = cell.value
                    
                    # Convert Python value to Lua representation
                    if value is None:
                        lua_value = "nil"
                    elif isinstance(value, bool):
                        lua_value = "true" if value else "false"
                    elif isinstance(value, (int, float)):
                        lua_value = str(value)
                    elif isinstance(value, str):
                        # Escape special characters for Lua strings
                        escaped = value.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")
                        lua_value = f'"{escaped}"'
                    else:
                        # For other types (datetime, etc.), convert to string
                        escaped = str(value).replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")
                        lua_value = f'"{escaped}"'
                    
                    lua_code.append(f'workbook["{sheet_name}"][{row_idx}][{col_idx}] = {lua_value}')
            
            lua_code.append("")  # Empty line between sheets
        
        # If no existing metadata, create a default metadata section
        if existing_metadata is None:
            existing_metadata = (
                "\n-- METADATA_START\n"
                "-- This section stores UI state and preferences\n"
                "local metadata = {\n"
                "    ColumnFilter = {}\n"
                "}\n"
                "\n"
                "return workbook, metadata"
            )
        else:
            # Use existing metadata as-is
            existing_metadata = "\n" + existing_metadata
        
        # Combine workbook data with metadata
        lua_code.append(existing_metadata)
        
        # Write to output file
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lua_code))
        
        print(f"SUCCESS: Converted to {output_path}")
        return 0
        
    except FileNotFoundError:
        print(f"ERROR: File not found: {excel_path}", file=sys.stderr)
        return 1
    except Exception as e:
        print(f"ERROR: {str(e)}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python excel_to_lua.py <input.xlsx> [output.lua]")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    output_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    sys.exit(excel_to_lua(excel_path, output_path))